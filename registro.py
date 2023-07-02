import datetime , pathlib , os
from openpyxl import Workbook,load_workbook
from openpyxl.styles import numbers


class Regitrador():
    def __init__(self,user:str) -> None:

        self.__user = user
        self.sheetName = f"{self.user}_moth_{datetime.date.today().month}"

        self.dataDeHoje = datetime.date(datetime.date.today().year,datetime.date.today().month,datetime.date.today().day)

        #control area
        self.EntradaStatus = False
        self.__DiaDeTrabalho = 0
        self.actualRow = 0

        #entrada e saida horarios 
        self.__HoraDiaEntrada = ""
        self.__HoraDiaSaida = ""
        
        self.__CreateSheet()
    

    
    @property
    def user(self):
        return self.__user
    
    @property
    def DiaDeTrabalho(self):
        return self.__DiaDeTrabalho
    
    @property
    def HoraDiaEntrada(self):
        return self.__HoraDiaEntrada
    
    @property
    def HoraDiaSaida(self):
        return self.__HoraDiaSaida
    
    
    
    def __VerifyArchive(self):
        try:
            path = pathlib.Path(f"{os.getcwd()}")
            userArchive = f"{self.user}.xlsx"
            arquivos = path.glob(f"**/{userArchive}")
            for x in arquivos:
                if userArchive in f"{x}".split("\\"):
                    return True
            return False

        except Exception as log:
            with open("Log_VerifyArchive.txt","w") as erro:
                erro.writelines(f"{log}")
    
    def __CreateSheet(self):
        auth = self.__VerifyArchive()
        if  auth == False:
            self.wb = Workbook()
            self.wb.create_sheet(f"{self.sheetName}",0)
            sheet = self.wb[f"{self.sheetName}"]
            sheet.append({"A":"Data de Entrada","B":"Hora de entrada","C":"Data de saida","D":"Hora de saida","E":"Horas trabalhadas"})
            self.wb.save(f"{self.user}.xlsx")

            return True
        
        else:
            return False
    
    

    def EntradaRegistro(self) ->dict:
        try:
            if self.EntradaStatus == False and self.__DiaDeTrabalho == 0:

                datatime = datetime.datetime.now()
            
                data  = self.dataDeHoje
                hora  = datatime.strftime("%H:%M")

                self.EntradaStatus = True
                self.__HoraDiaEntrada = {"entrada":{"data":data,"hora":hora},"tipo":"entrada"}

                wb =  load_workbook(f"{os.getcwd()}\{self.user}.xlsx")
                if not self.sheetName in wb.sheetnames:
                    wb.create_sheet(f"{self.sheetName}")
                sheet = wb[f"{self.sheetName}"]
                actualRow = sheet.max_row+1
                sheet.cell(row=actualRow,column=1,value=data)
                sheet.cell(row=actualRow,column=2,value=hora)
                column1 = sheet["A"]
                column2 = sheet["C"]
                for x in column1:
                    x.number_format = f"{numbers.FORMAT_DATE_DDMMYY}"
                for x in column2:
                    x.number_format = f"{numbers.FORMAT_DATE_DDMMYY}"

                wb.save(f"{self.user}.xlsx")

                return self.HoraDiaEntrada
            else:
                print(f"User : {self.user} Já esta trabalhando.")
                return False
            
        except Exception as log:
            with open("Log_EntradaRegistro.txt","w") as erro:
                erro.writelines(f"{log}")

    def SaidaRegistro(self) -> dict:
        try:               
            datatime = datetime.datetime.now()
            
            data  = self.dataDeHoje
            hora  = datatime.strftime("%H:%M")

            self.__HoraDiaSaida = {"saida":{"data":data,"hora":hora},"tipo":"saida"}
            self.EntradaStatus = False
            self.__DiaDeTrabalho = 1
            

            wb =  load_workbook(f"{os.getcwd()}\{self.user}.xlsx")
            sheet = wb[f"{self.sheetName}"]
            actualRow = sheet.max_row
            sheet.cell(row=actualRow,column=3,value=data)
            sheet.cell(row=actualRow,column=4,value=hora)
            
            
            column1 = sheet["A"]
            column2 = sheet["C"]
            for x in column1:
                x.number_format = f"{numbers.FORMAT_DATE_DDMMYY}"
            for x in column2:
                x.number_format = f"{numbers.FORMAT_DATE_DDMMYY}"
                

            wb.save(f"{self.user}.xlsx")
            
            wb =  load_workbook(f"{os.getcwd()}\{self.user}.xlsx")
            sheet = wb[f"{self.sheetName}"]
            sheet.cell(row=actualRow,column=5,value=self.HorasTrabalhadas())
            wb.save(f"{self.user}.xlsx")

            return self.HoraDiaSaida
            
        except Exception as log:
            with open("Log_SaidaRegistro.txt","w") as erro:
                erro.writelines(f"{log}")

    def HorasTrabalhadas(self):
        try:
            wb =  load_workbook(f"{os.getcwd()}\{self.user}.xlsx")
            sheet = wb[f"{self.sheetName}"]
            maxrow = self.actualRow = sheet.max_row

            timeconvert = lambda x: datetime.datetime.strptime(str(x),"%H:%M")

            horaEntrada = timeconvert(sheet.cell(row=maxrow,column=2).value)

            horaSaida = timeconvert(sheet.cell(row=maxrow,column=4).value)

            horasTrabalhadas = (horaSaida - horaEntrada) - datetime.timedelta(minutes=30)      

            wb.close()
            return  horasTrabalhadas 

        except Exception as log:
            with open("Log_HorasTrabalhadas.txt","w") as erro:
               erro.writelines(f"{log}")